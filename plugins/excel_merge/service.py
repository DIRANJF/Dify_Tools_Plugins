"""
Excel 合并工具 — 核心业务逻辑
"""

import io
import re
from copy import copy
from typing import Optional

import openpyxl
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def download_excel(url: str) -> openpyxl.Workbook:
    """从 URL 下载 Excel 文件并返回 Workbook 对象"""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)


def get_first_sheet(wb: openpyxl.Workbook):
    return wb.worksheets[0]


def extract_headers(sheet, header_rows: int = 2):
    """
    提取表头信息（默认前两行）。
    返回: (row1, row2, merged_labels)
    """
    row1 = []
    row2 = []
    for col in range(1, sheet.max_column + 1):
        v1 = sheet.cell(row=1, column=col).value
        row1.append(str(v1).strip() if v1 is not None else "")
    if header_rows >= 2 and sheet.max_row >= 2:
        for col in range(1, sheet.max_column + 1):
            v2 = sheet.cell(row=2, column=col).value
            row2.append(str(v2).strip() if v2 is not None else "")

    merged_labels = []
    for i in range(len(row1)):
        r2 = row2[i] if i < len(row2) else ""
        r1 = row1[i]
        merged_labels.append(r2 if (r2 and r2 != r1) else r1)

    return row1, row2, merged_labels


def is_total_row(row_data: list, keywords: list[str] = None) -> bool:
    """判断某一行是否为"合计"行"""
    if keywords is None:
        keywords = ["合计", "总计", "小计", "total", "sum", "Total", "SUM"]
    for val in row_data:
        if val is None:
            continue
        s = str(val).strip()
        if any(kw in s for kw in keywords):
            return True
    return False


def find_total_columns(header_labels: list, keywords: list[str] = None) -> list[int]:
    """查找表头中属于"合计"类的列索引（0-based）"""
    if keywords is None:
        keywords = ["合计", "总计", "小计", "total", "sum"]
    result = []
    for idx, label in enumerate(header_labels):
        if any(kw in label for kw in keywords):
            result.append(idx)
    return result


def build_column_mapping(labels_f1: list, labels_f2: list) -> dict[int, int]:
    """
    构建 file2 列 → file1 列的映射。
    匹配策略：精确 → 包含 → 模糊（去特殊字符）。
    返回: {file2_col_idx: file1_col_idx}
    """
    def normalize(s: str) -> str:
        s = s.strip().lower()
        return re.sub(r'[\s\u3000\n\r\t()（）【】\[\]{}·・\-_/／\\]', '', s)

    norm1 = [normalize(h) for h in labels_f1]
    norm2 = [normalize(h) for h in labels_f2]
    used: set[int] = set()
    mapping: dict[int, int] = {}

    # 第一轮：精确匹配
    for f2i, h2 in enumerate(norm2):
        if not h2:
            continue
        for f1i, h1 in enumerate(norm1):
            if f1i in used:
                continue
            if h2 == h1:
                mapping[f2i] = f1i
                used.add(f1i)
                break

    # 第二轮：包含匹配
    for f2i, h2 in enumerate(norm2):
        if f2i in mapping or not h2:
            continue
        best, best_len = None, 0
        for f1i, h1 in enumerate(norm1):
            if f1i in used or not h1:
                continue
            if h2 in h1 or h1 in h2:
                overlap = min(len(h2), len(h1))
                if overlap > best_len:
                    best, best_len = f1i, overlap
        if best is not None:
            mapping[f2i] = best
            used.add(best)

    return mapping


def copy_cell_style(src, dst):
    """复制单元格样式"""
    try:
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.number_format:
            dst.number_format = src.number_format
    except Exception:
        pass


def detect_numeric_columns(rows: list, labels: list) -> list[int]:
    """检测哪些列是数值列"""
    if not rows:
        return []
    candidates = set(range(len(labels)))
    non_numeric: set[int] = set()
    for row in rows:
        for ci in candidates:
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None or v == "":
                continue
            if not isinstance(v, (int, float)):
                try:
                    float(str(v).strip())
                except (ValueError, TypeError):
                    non_numeric.add(ci)
    return sorted(candidates - non_numeric)


# ---------------------------------------------------------------------------
# 核心合并逻辑
# ---------------------------------------------------------------------------

def merge_excel_files(file1_url: str, file2_url: str) -> bytes:
    """
    1. 下载并解析两个 Excel 文件
    2. 从文件1提取表头（前两行）
    3. 从文件2提取表头和数据
    4. 匹配字段，将文件2数据填充到文件1表头下
    5. 生成合计行
    返回: 生成的 Excel 文件 bytes
    """
    wb1 = download_excel(file1_url)
    wb2 = download_excel(file2_url)
    ws1 = get_first_sheet(wb1)
    ws2 = get_first_sheet(wb2)

    # --- 文件1 表头 ---
    h_row1, h_row2, h_labels1 = extract_headers(ws1, header_rows=2)
    num_cols = len(h_labels1)

    # --- 文件1 原有数据行（第3行起），识别合计行 ---
    f1_data_rows: list[list] = []
    f1_total_indices: list[int] = []
    for r in range(3, ws1.max_row + 1):
        row = [ws1.cell(r, c).value for c in range(1, ws1.max_column + 1)]
        if is_total_row(row):
            f1_total_indices.append(len(f1_data_rows))
        f1_data_rows.append(row)

    # --- 文件2 表头 & 数据 ---
    f2_header = [
        str(ws2.cell(1, c).value).strip() if ws2.cell(1, c).value is not None else ""
        for c in range(1, ws2.max_column + 1)
    ]
    f2_data_rows: list[list] = []
    for r in range(2, ws2.max_row + 1):
        row = [ws2.cell(r, c).value for c in range(1, ws2.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        f2_data_rows.append(row)

    # --- 列映射 ---
    col_map = build_column_mapping(h_labels1, f2_header)

    # --- 新建工作簿 ---
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = ws1.title or "Sheet1"

    # 写入表头（前两行）并复制样式
    for ci in range(1, num_cols + 1):
        src1 = ws1.cell(1, ci)
        dst1 = ws_new.cell(1, ci)
        dst1.value = src1.value
        copy_cell_style(src1, dst1)
        if ws1.max_row >= 2:
            src2 = ws1.cell(2, ci)
            dst2 = ws_new.cell(2, ci)
            dst2.value = src2.value
            copy_cell_style(src2, dst2)

    # 复制列宽
    for ci in range(1, num_cols + 1):
        letter = get_column_letter(ci)
        w = ws1.column_dimensions[letter].width
        if w:
            ws_new.column_dimensions[letter].width = w

    # --- 填充数据 ---
    start_row = 3
    filled: list[list] = []
    for i, f2_row in enumerate(f2_data_rows):
        out_row = [None] * num_cols
        for f2_ci, f1_ci in col_map.items():
            if f2_ci < len(f2_row):
                val = f2_row[f2_ci]
                out_row[f1_ci] = val
                cell = ws_new.cell(start_row + i, f1_ci + 1)
                cell.value = val
                # 复制文件1第3行样式
                if ws1.max_row >= 3:
                    try:
                        copy_cell_style(ws1.cell(3, f1_ci + 1), cell)
                    except Exception:
                        pass
        filled.append(out_row)

    # --- 合计行 ---
    total_cols = find_total_columns(h_labels1)
    numeric_cols = detect_numeric_columns(filled, h_labels1)
    sum_cols = [c for c in total_cols if c in numeric_cols] or numeric_cols

    has_total_tpl = len(f1_total_indices) > 0

    if filled and sum_cols:
        tr_idx = start_row + len(filled)
        # 标签列
        label_written = False
        for ci in range(num_cols):
            if ci not in sum_cols and not label_written:
                ws_new.cell(tr_idx, ci + 1).value = "合计"
                label_written = True
                break
        if not label_written:
            ws_new.cell(tr_idx, 1).value = "合计"

        for ci in sum_cols:
            total = 0.0
            for row in filled:
                if ci < len(row) and row[ci] is not None:
                    try:
                        total += float(row[ci])
                    except (ValueError, TypeError):
                        pass
            cell = ws_new.cell(tr_idx, ci + 1)
            cell.value = int(total) if total == int(total) else round(total, 2)
            # 复制合计模板样式
            if has_total_tpl:
                try:
                    copy_cell_style(
                        ws1.cell(f1_total_indices[0] + 3, ci + 1), cell
                    )
                except Exception:
                    pass
            # 加粗
            try:
                cell.font = Font(
                    name=cell.font.name if cell.font else None,
                    size=cell.font.size if cell.font else None,
                    bold=True,
                    color=cell.font.color if cell.font else None,
                )
            except Exception:
                pass

    # --- 复制表头合并单元格 ---
    for merged in ws1.merged_cells.ranges:
        if merged.min_row <= 2:
            try:
                ws_new.merge_cells(str(merged))
            except Exception:
                pass

    buf = io.BytesIO()
    wb_new.save(buf)
    buf.seek(0)
    return buf.getvalue()
